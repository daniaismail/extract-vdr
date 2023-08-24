import pandas as pd
from openpyxl import load_workbook

file = 'vdr/Jadestone VDR - 24.08.2023.xlsx'

wb = load_workbook(file, data_only=True)
sheet = wb['VDR']

vesselName = sheet['E12'].value
captainName = sheet['J12'].value
location = sheet['L12'].value
engineHour = sheet['J23:J44']
engineFuel = sheet['T23:T45']
weather = sheet['I16:O19']
activity = sheet['I76':'I87']
weather_data = []

for row in sheet.iter_rows(min_row=16,max_row=19,min_col=9,max_col=15):
    weather_data.append([cell.value for cell in row])

engineHour_value = [cell[0].value for cell in engineHour]
engineFuel_value = [cell[0].value for cell in engineFuel]
activity_value = [cell[0].value for cell in activity]

engine_hour_column = ['PORT ME (hrs)','PORT ME OUTLET FLOWMETER (hrs)','CENTER ME (P) (hrs)','CENTER ME (P) OUTLET FLOWMETER (hrs)','CENTER ME (STBD) (hrs)','CENTER ME (STBD) OUTLET FLOWMETER (hrs)','STBD ME (hrs)','STBD ME OUTLET FLOWMETER (hrs)','BOW THRUSTER 1 (hrs)','BOW THRUSTER 2 (hrs)','BOW THRUSTER 3 (hrs)','STERN THRUSTER 1 (hrs)','STERN THRUSTER 2 (hrs)','SHAFT GENERATOR 1 (hrs)','SHAFT GENERATOR 2 (hrs)','GENERATOR 1 (hrs)','GENERATOR 2 (hrs)','GENERATOR 3 (hrs)','GENERATOR 4 (hrs)','GENERATOR 5 (hrs)','GENERATOR 6 (hrs)','EMERGENCY GENERATOR (hrs)']
engine_fuel_column = ['PORT ME (ltrs)','PORT ME OUTLET FLOWMETER (ltrs)','CENTER ME (P) (ltrs)','CENTER ME (P) OUTLET FLOWMETER (ltrs)','CENTER ME (STBD) (ltrs)','CENTER ME (STBD) OUTLET FLOWMETER (ltrs)','STBD ME (ltrs)','STBD ME OUTLET FLOWMETER (ltrs)','BOW THRUSTER 1 (ltrs)','BOW THRUSTER 2 (ltrs)','BOW THRUSTER 3 (ltrs)','STERN THRUSTER 1 (ltrs)','STERN THRUSTER 2 (ltrs)','SHAFT GENERATOR 1 (ltrs)','SHAFT GENERATOR 2 (ltrs)','GENERATOR 1 (ltrs)','GENERATOR 2 (ltrs)','GENERATOR 3 (ltrs)','GENERATOR 4 (ltrs)','GENERATOR 5 (ltrs)','GENERATOR 6 (ltrs)','EMERGENCY GENERATOR (ltrs)','fuel']
weather_column = ['TIME','WIND','SWELL','SEA','SKY','VISIBILITY','TEMP ( °C )']
activity_column = ['Maneuvering at Supply Base','Alongside berth at Supply Base','Anchorage at Supply Base','-','En-route: full speed','En-route: economical speed','-','Inter-rig/ Maneuvering offshore','Standby steaming offshore','Cargo works within 500m zone','Towing/ Static Towing/ Rigmove/ Hose Handling','Mooring to buoy/platform offshore']

df = pd.DataFrame({'Vessel Name': [vesselName],
                   'Captain on Duty': [captainName],
                   'Location @ 2400H': [location]})
df_engine_hour = pd.DataFrame([engineHour_value], columns=engine_hour_column)
df_engine_fuel = pd.DataFrame([engineFuel_value], columns=engine_fuel_column)
df_weather = pd.DataFrame(weather_data,columns=weather_column)
df_activity= pd.DataFrame([activity_value],columns=activity_column)

df_summary = [df,df_engine_hour,df_engine_fuel]
combined_1 = pd.concat(df_summary, axis=1)
concat_both = combined_1.reset_index(drop=True)
df_weather = df_weather.reset_index(drop=True)
df_activity = df_activity.reset_index(drop=True)

with pd.ExcelWriter('AISHAH AIMS 3.xlsx', engine = 'openpyxl', mode='a') as writer:
    concat_both.to_excel(writer, sheet_name='Summary', index=False)
    df_weather.to_excel(writer, sheet_name='Weather', index=False)
    df_activity.to_excel(writer, sheet_name='Activity', index=False)

#combined_1.to_excel(r'AISHAH AIMS 3.xlsx', sheet_name='Summary', index=False)
#weather_table.to_excel(r'AISHAH AIMS 3.xlsx', sheet_name='Weather', index=False)
memory_usage_bytes = df.memory_usage(deep=True).sum()
print(memory_usage_bytes)
