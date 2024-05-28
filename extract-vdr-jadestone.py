import imaplib
import email
import sys
import pandas as pd
import os
import shutil
from openpyxl import load_workbook
from datetime import date, timedelta
import logging

vdrDirectory = 'VDR'
vdrClientDir = 'C:/Users/User/PycharmProjects/extract-vdr/vdr/jadestone'

date_date = date.today()
t = date_date.strftime("%d%m%y")
logging.basicConfig(format='%(asctime)s %(message)s', datefmt='%d%m%Y %I:%M:%S%p', filename=f"C:/Users/User/PycharmProjects/extract-vdr/log/log {t}.txt", level=logging.DEBUG)
yesterday = date_date - timedelta(days=1)
todayDate = yesterday.strftime("%d/%m/%Y")
dt = date_date.strftime('%d-%b-%Y')
fileTemplate = 'Template.xlsx'

#EMAIL INFO
email_vdr = "mvmcc@meridiansurveys.com.my"
pwd_vdr = "dc)in]}Xzk&%"
server_mssb = "meridian-svr.meridiansurveys.com.my"

#READ FROM TXT FILE AND APPEND INTO LIST
vesselemail_list = []

logging.debug('Logging started..')
logging.debug('Open jadestone-email-list.txt')
try:
    with open(r'C:\Users\User\PycharmProjects\extract-vdr\jadestone-email-list.txt') as f:
        try:
            for line in f:
                vesselemail_list.append(line.replace("\n", ""))
        except:
            logging.error('Error info in text file. Check for typo')
except:
    logging.exception('Error info in text file or check txt file name')
    sys.exit()

def delete_files_in_directory(vdrClientDir):
   try:
     files = os.listdir(vdrClientDir)
     for file in files:
       file_path = os.path.join(vdrClientDir, file)
       if os.path.isfile(file_path):
         os.remove(file_path)
     print("All files deleted successfully.")
   except OSError:
     print("Error occurred while deleting files.")

#DOWNLOAD FROM EMAIL
def dwl_vdr(email_add, password, server, vesselEmail):
    imap = imaplib.IMAP4_SSL(server, 993)
    imap.login(email_add, password)
    imap.select('INBOX')

    index = 0

    while index < len(vesselEmail):
        try:
            logging.debug(f'Download from email {vesselEmail[index]}')
            typ, data = imap.search(None, '(SINCE %s)' % (dt,),'(FROM %s)' % (vesselEmail[index],))

            print(vesselEmail[index])
            for num in data[0].split():
                typ, data = imap.fetch(num, '(RFC822)')
                raw_email = data[0][1]
                raw_email_string = raw_email.decode('ISO-8859–1')
                email_message = email.message_from_string(raw_email_string)
                subject_name = email_message['subject']
                print(subject_name)

                # att_path = "No attachment found from email " + subject_name
                logging.debug(f'Email subject: {subject_name}')
                for part in email_message.walk():
                    if part.get_content_maintype() == 'multipart':
                        continue
                    if part.get('Content-Disposition') is None:
                        continue

                    if part.get_filename() and any(keyword in part.get_filename() for keyword in ['VDR','VDMR']) and part.get_filename().endswith('.xlsx'):
                        filename = os.path.join(vdrClientDir, part.get_filename())
                        logging.debug(f'file path: {filename}')
                        print(filename)
                        with open(filename, 'wb') as f:
                            f.write(part.get_payload(decode=True))
                        logging.debug(f'File downloaded: {filename}')
            index += 1

        except (OSError, TypeError) as k:
            print(k)
            continue

    imap.close()
    imap.logout()
    logging.debug('Job finished..')

try:
    logging.debug('Deleting old vdr files and Downloading from mvmcc@meridiansurveys.com.my..')
    delete_files_in_directory(vdrClientDir)
    dwl_vdr(email_vdr, pwd_vdr, server_mssb, vesselemail_list)

except Exception as e:
    logging.error("Error occurred", exc_info=True)

#START EXTRACT
for vdrFile in os.listdir(vdrClientDir):
    if vdrFile.endswith('.xlsx'):
        vdrFileDir = os.path.join(vdrClientDir,vdrFile)
        wb = load_workbook(vdrFileDir, data_only=True)
        sheetVDR = wb['VDR']
        vesselName = sheetVDR['E12'].value
        fileName = vesselName + '.xlsx'
        #DUPLICATE TEMPLATE
        template_path = os.path.dirname(fileTemplate)
        file_path = os.path.join(template_path,fileName)
        shutil.copy(fileTemplate,file_path)
        sheetCrew = wb['CORE-CREW']
        #shipMMSI = 533180042
        #shipNationality = 'MALAYSIA'
        #shipType = 'UV'
        captainName = sheetVDR['J12'].value
        location = sheetVDR['L12'].value
        engineHour = sheetVDR['J23:J44']
        engineFuel = sheetVDR['T23:T45']
        activity = sheetVDR['I76':'I87']
        weather_data = []
        activityLog_data = []
        crew_data = []
        rob_data = []

        #GET DATA
        engineHour_value = [cell[0].value for cell in engineHour]
        engineFuel_value = [cell[0].value for cell in engineFuel]
        activity_value = [cell[0].value for cell in activity]
        for row in sheetVDR.iter_rows(min_row=14,max_row=103,min_col=23,max_col=35):
            activityLog_data.append([cell.value for cell in row])
        for row in sheetVDR.iter_rows(min_row=16,max_row=19,min_col=9,max_col=15):
            weather_data.append([cell.value for cell in row])
        for row in sheetCrew.iter_rows(min_row=6,max_row=30,min_col=2,max_col=9):
            crew_data.append([cell.value for cell in row])
        for row in sheetVDR.iter_rows(min_row=52,max_row=72,min_col=2,max_col=21):
            rob_data.append([cell.value for cell in row])

        #COLUMN NAME
        engine_hour_column = ['PORT ME (hrs)','PORT ME OUTLET FLOWMETER (hrs)','CENTER ME (P) (hrs)','CENTER ME (P) OUTLET FLOWMETER (hrs)','CENTER ME (STBD) (hrs)','CENTER ME (STBD) OUTLET FLOWMETER (hrs)','STBD ME (hrs)','STBD ME OUTLET FLOWMETER (hrs)','BOW THRUSTER 1 (hrs)','BOW THRUSTER 2 (hrs)','BOW THRUSTER 3 (hrs)','STERN THRUSTER 1 (hrs)','STERN THRUSTER 2 (hrs)','SHAFT GENERATOR 1 (hrs)','SHAFT GENERATOR 2 (hrs)','GENERATOR 1 (hrs)','GENERATOR 2 (hrs)','GENERATOR 3 (hrs)','GENERATOR 4 (hrs)','GENERATOR 5 (hrs)','GENERATOR 6 (hrs)','EMERGENCY GENERATOR (hrs)']
        engine_fuel_column = ['PORT ME (ltrs)','PORT ME OUTLET FLOWMETER (ltrs)','CENTER ME (P) (ltrs)','CENTER ME (P) OUTLET FLOWMETER (ltrs)','CENTER ME (STBD) (ltrs)','CENTER ME (STBD) OUTLET FLOWMETER (ltrs)','STBD ME (ltrs)','STBD ME OUTLET FLOWMETER (ltrs)','BOW THRUSTER 1 (ltrs)','BOW THRUSTER 2 (ltrs)','BOW THRUSTER 3 (ltrs)','STERN THRUSTER 1 (ltrs)','STERN THRUSTER 2 (ltrs)','SHAFT GENERATOR 1 (ltrs)','SHAFT GENERATOR 2 (ltrs)','GENERATOR 1 (ltrs)','GENERATOR 2 (ltrs)','GENERATOR 3 (ltrs)','GENERATOR 4 (ltrs)','GENERATOR 5 (ltrs)','GENERATOR 6 (ltrs)','EMERGENCY GENERATOR (ltrs)','fuel']
        activity_column = ['Maneuvering at Supply Base','Alongside berth at Supply Base','Anchorage at Supply Base','','En-route: full speed','En-route: economical speed','','Inter-rig/ Maneuvering offshore','Standby steaming offshore','Cargo works within 500m zone','Towing/ Static Towing/ Rigmove/ Hose Handling','Mooring to buoy/platform offshore']
        weather_column = ['TIME','WIND','SWELL','SEA','SKY','VISIBILITY','TEMP ( °C )']
        activityLog_column = ['FROM (0000)','TO (2400)','DURATION','ACTIVITY & LOCATION (Consecutive entries - no gaps allowed)','','','','','VESSEL MOVEMENT CATEGORY','','LOCATION','DP MODE (Y/N)','NON-CREW POB']
        crew_column = ['No.','Name','Rank','Age','Nationality','IC or Passport No.','Date Sign-on(DD/MM/YYYY)','No. of working days (max. 90 days)']
        rob_column = ['ROB','OPENING @ 0000H','','','','','','Consumed','','','','','','','','','','','CLOSING @ 2400H','Remarks']

        #CREATE DATAFRAME
        df = pd.DataFrame({#'MMSI': shipMMSI,
                           'Vessel Name': [vesselName],
                           #'Nationality of Ship': shipNationality,
                           #'Type of Vessel': shipType,
                           'Captain on Duty': [captainName],
                           'Location @ 2400H': [location],
                           'date': todayDate})
        df_engine_hour = pd.DataFrame([engineHour_value], columns=engine_hour_column)
        df_engine_fuel = pd.DataFrame([engineFuel_value], columns=engine_fuel_column)
        df_activity= pd.DataFrame([activity_value],columns=activity_column)
        df_activityLog = pd.DataFrame(activityLog_data,columns=activityLog_column)
        df_weather = pd.DataFrame(weather_data,columns=weather_column)
        df_crew = pd.DataFrame(crew_data,columns=crew_column)
        df_rob = pd.DataFrame(rob_data,columns=rob_column)

        df_summary = [df,df_engine_hour,df_engine_fuel]
        combined_1 = pd.concat(df_summary, axis=1)

        #RESET INDEX
        concat_both = combined_1.reset_index(drop=True)
        df_weather = df_weather.reset_index(drop=True)
        df_activity = df_activity.reset_index(drop=True)
        df_activityLog = df_activityLog.reset_index(drop=True)
        df_crew = df_crew.reset_index(drop=True)
        df_rob = df_rob.reset_index(drop=True)

        #EXPORT TO_EXCEL
        with pd.ExcelWriter(fileName, engine = 'openpyxl', mode='a') as writer:
            concat_both.to_excel(writer, sheet_name='Summary', index=False)
            df_weather.to_excel(writer, sheet_name='Weather', index=False)
            df_activity.to_excel(writer, sheet_name='Activity', index=False)
            df_activityLog.to_excel(writer, sheet_name='Activity log', index=False)
            df_crew.to_excel(writer, sheet_name='Crew list', index=False)
            df_rob.to_excel(writer, sheet_name='ROB', index=False)

#combined_1.to_excel(r'AISHAH AIMS 3.xlsx', sheet_name='Summary', index=False)
#weather_table.to_excel(r'AISHAH AIMS 3.xlsx', sheet_name='Weather', index=False)
#memory_usage_bytes = df_activityLog.memory_usage(deep=True).sum()
#print(memory_usage_bytes)