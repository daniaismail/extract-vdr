import sys
import imaplib
import email
import sys
import pandas as pd
import os
import shutil
from openpyxl import load_workbook
from datetime import date, timedelta
import logging
from email.header import decode_header
from tnefparse import TNEF

#VDR DIRECTORY
vdrClientDir = 'C:/Users/User/PycharmProjects/extract-vdr/vdr/enquest'

#FUNCTIONS
#download vdr from email

def delete_files_in_directory(vdrClientDir):
   try:
     files = os.listdir(vdrClientDir)
     for file in files:
       file_path = os.path.join(vdrClientDir, file)
       if os.path.isfile(file_path):
         os.remove(file_path)
     print("All files deleted successfully.")
   except OSError:
     print("Error occurred while deleting files.")

def dwl_vdr(email_add, password, server, vesselEmail):
    imap = imaplib.IMAP4_SSL(server, 993)
    imap.login(email_add, password)
    imap.select('INBOX')

    index = 0

    while index < len(vesselEmail):
        try:
            logging.debug(f'Download from email {vesselEmail[index]}')
            typ, data = imap.search(None, '(SINCE %s)' % (dt,),'(FROM %s)' % (vesselEmail[index],))

            print(vesselEmail[index])
            for num in data[0].split():
                typ, data = imap.fetch(num, '(RFC822)')
                raw_email = data[0][1]

                '''
                raw_email_string = raw_email.decode('ISO-8859–1')
                email_message = email.message_from_string(raw_email_string)
                subject_name = email_message['subject']
                print(subject_name)
                '''
                email_message = email.message_from_bytes(raw_email)
                subject, encoding = decode_header(email_message["Subject"])[0]

                if isinstance(subject, bytes):
                    subject = subject.decode(encoding or "UTF-8", errors="replace")

                print(f'Subject: {subject}')

                # att_path = "No attachment found from email " + subject_name
                logging.debug(f'Email subject: {subject}')
                for part in email_message.walk():
                    try:
                        if part.get_content_maintype() == 'multipart':
                            continue
                        if part.get('Content-Disposition') is None:
                            continue

                        fileName = part.get_filename()

                        if fileName:
                            decode_name, encoding = decode_header(fileName)[0]
                            if isinstance(decode_name, bytes):
                                decode_name = decode_name.decode(encoding or "UTF-8", errors="replace")

                        if decode_name and any(keyword in decode_name for keyword in ['VDR','VDMR']) and decode_name.endswith('.xlsx'):
                            try:
                                att_path = os.path.join(vdrClientDir, decode_name)
                                print(decode_name)
                                print(f'Path: {att_path}')

                                if not os.path.isfile(att_path):
                                    with open(att_path, "wb") as fp:
                                        fp.write(part.get_payload(decode=True))
                            except TypeError as e:
                                continue

                        if decode_name and any(keyword in decode_name for keyword in ['winmail']) and decode_name.endswith('.dat'):
                            try:
                                att_path = os.path.join(vdrClientDir, decode_name)
                                print(decode_name)
                                print(f'Path: {att_path}')

                                if not os.path.isfile(att_path):
                                    with open(att_path, "wb") as fp:
                                        fp.write(part.get_payload(decode=True))
                                    extract_winmail_dat("C:/Users/User/PycharmProjects/extract-vdr/vdr/enquest/winmail.dat", "C:/Users/User/PycharmProjects/extract-vdr/vdr/enquest")
                                    os.remove(att_path)
                                    print("Remove dat file")
                            except TypeError as e:
                                continue

                    except Exception as e:
                        print(f"Error processing attachment: {e}")


                '''
                                for part in email_message.walk():
                    if part.get_content_maintype() == 'multipart':
                        continue
                    if part.get('Content-Disposition') is None:
                        continue

                    if part.get_filename() and any(keyword in part.get_filename() for keyword in ['VDR','VDMR']) and part.get_filename().endswith('.xlsx'):
                        filename = os.path.join(vdrClientDir, part.get_filename())
                        logging.debug(f'file path: {filename}')
                        print(filename)
                        with open(filename, 'wb') as f:
                            f.write(part.get_payload(decode=True))
                        logging.debug(f'File downloaded: {filename}')
                '''

            index += 1

        except (OSError, TypeError) as k:
            print(k)
            continue

    imap.close()
    imap.logout()
    logging.debug('Job finished..')

def extract_winmail_dat(file_path, output_directory="."):
    """
    Extracts attachments and body from a winmail.dat file.

    Args:
        file_path (str): The path to the winmail.dat file.
        output_directory (str): The directory where extracted files will be saved.
    """
    try:
        with open(file_path, 'rb') as f:
            winmail_data = f.read()

        tnef_data = TNEF(winmail_data)

        # Extract attachments
        for attachment in tnef_data.attachments:
            attachment_name = attachment.long_filename() or attachment.name
            with open(f"{output_directory}/{attachment_name}", 'wb') as out_f:
                out_f.write(attachment.data)
            print(f"Extracted attachment: {attachment_name}")

        # Extract email body (if present)
        if tnef_data.body:
            body_content = tnef_data.body.decode('utf-8', errors='ignore')
            with open(f"{output_directory}/email_body.txt", 'w', encoding='utf-8') as body_f:
                body_f.write(body_content)
            print("Extracted email body.")

    except Exception as e:
        print(f"Error extracting winmail.dat: {e}")

def normalize_sheet_name(sheet_name):
    return sheet_name.lower().strip()

#EMAIL INFO
email_vdr = "mvmcc@meridiansurveys.com.my"
pwd_vdr = "HehAb58ynLR3FX"
server_mssb = "meridian-svr.meridiansurveys.com.my"

#READ FROM TXT FILE AND APPEND INTO LIST
vesselemail_list = []

date_date = date.today()
dt = date_date.strftime('%d-%b-%Y')
yesterday = date_date - timedelta(days=1)
day_yesterday = yesterday.day
todayDate = yesterday.strftime("%d/%m/%Y")
fileTemplate = 'Template.xlsx'

logging.debug('Logging started..')
logging.debug('Open enquest-email-list.txt')
try:
    with open(r'C:\Users\User\PycharmProjects\extract-vdr\enquest-email-list.txt') as f:
        try:
            for line in f:
                vesselemail_list.append(line.replace("\n", ""))
        except:
            logging.error('Error info in text file. Check for typo')
except:
    logging.exception('Error info in text file or check txt file name')
    sys.exit()

try:
    logging.debug('Deleting old vdr files and Downloading from mvmcc@meridiansurveys.com.my..')
    delete_files_in_directory(vdrClientDir)
    dwl_vdr(email_vdr, pwd_vdr, server_mssb, vesselemail_list)

except Exception as e:
    logging.error("Error occurred", exc_info=True)

for vdrFile in os.listdir(vdrClientDir):
    if vdrFile.endswith('.xlsx'):
        print(f'Accessing: {vdrFile}')
        vdrFileDir = os.path.join(vdrClientDir, vdrFile)
        wb = load_workbook(vdrFileDir, data_only=True)

        # Normalize sheet names
        normalized_sheet_names = [normalize_sheet_name(sheet_name) for sheet_name in wb.sheetnames]

        sheetDailyReport = None
        sheetDailyReport2 = None
        sheetActivities = None
        sheetFuel = None
        sheetTOD = None

        # Find the sheets dynamically
        for normalized_name, sheet_name in zip(normalized_sheet_names, wb.sheetnames):
            if normalized_name == 'daily report':
                sheetDailyReport = wb[sheet_name]
            elif normalized_name == 'daily report(2)':
                sheetDailyReport2 = wb[sheet_name]
            elif normalized_name == 'daily report3':
                sheetDailyReport2 = wb[sheet_name]
            elif normalized_name == 'daily report2':
                sheetDailyReport2 = wb[sheet_name]
            elif normalized_name == 'boat movements':
                sheetActivities = wb[sheet_name]
            elif normalized_name == 'fuel monitoring':
                sheetFuel = wb[sheet_name]
            elif normalized_name == 'tod report':
                sheetTOD = wb[sheet_name]

        if sheetDailyReport is None or sheetDailyReport2 is None or sheetActivities is None or sheetFuel is None or sheetTOD is None:
            print(f'Error: One or more required sheets not found in {vdrFile}, sheet {sheet_name}')
            continue

        vesselName = sheetDailyReport['C4'].value
        fileName = 'FROM VDR ' + vesselName + '.xlsx'
        # DUPLICATE TEMPLATE
        template_path = os.path.dirname(fileTemplate)
        file_path = os.path.join(template_path, fileName)
        shutil.copy(fileTemplate, file_path)
        location = sheetDailyReport['U4'].value
        enroute_to = sheetDailyReport['W4'].value
        weather_data_1 = []
        weather_data_2 = []
        rob_data = []
        fuel_data = []
        activities_data = []
        operation_act_p1_data = []
        operation_act_p2_data = []
        TOD_data = []
        cargo_descriptions = ['FUEL OIL (Ltrs)', 'FRESH WATER (Ltrs)', 'LUB OIL (Ltrs)', 'DISPERSANT (Ltrs)']

        #GET DATA
        for row in sheetActivities.iter_rows(min_row=5,max_row=35,min_col=6,max_col=17):
            activities_data.append([cell.value for cell in row])
        for row in sheetFuel.iter_rows(min_row=8,max_row=38,min_col=4,max_col=22):
            fuel_data.append([cell.value for cell in row])
        for row in sheetDailyReport.iter_rows(min_row=7,max_row=9,min_col=3,max_col=4):
            weather_data_1.append([cell.value for cell in row])
        for row in sheetDailyReport.iter_rows(min_row=7,max_row=9,min_col=14,max_col=15):
            weather_data_2.append([cell.value for cell in row])
        for row in sheetDailyReport.iter_rows(min_row=13,max_row=16,min_col=14,max_col=28):
            rob_data.append([cell.value for cell in row])
        for row in sheetTOD.iter_rows(min_row=18,max_row=40,min_col=3,max_col=19):
            TOD_data.append([cell.value for cell in row])
        for row in sheetDailyReport.iter_rows(min_row=66,max_row=150,min_col=3,max_col=21):
            operation_act_p1_data.append([cell.value for cell in row])
        for row in sheetDailyReport2.iter_rows(min_row=26,max_row=50,min_col=3,max_col=15):
            operation_act_p2_data.append([cell.value for cell in row])

        #COLUMN NAME
        summary_column = ['MMSI','Vessel Name','Nationality of Ship','Type of Vessel','Location','Date']
        weather_column = ['column-1', 'column-2']
        fuel_column = ['Port ME (HRS)','Port ME (LTRS)','Centre ME (HRS)','Centre ME (LTRS)','Stbd ME (HRS)','Stbd ME (LTRS)','Genset 1 (HRS)','Genset 1 (LTRS)','Genset 2 (HRS)','Genset 2 (LTRS)','Genset 3 (HRS)','Genset 3 (LTRS)','Bow Thruster (HRS)','Bow Thruster (LTRS)','Others (HRS)','Others (LTRS)','Main Eng. Cons. (LTRS)','Aux Eng. Cons. (LTRS)','Total Daily Cons. (LTRS)']
        activities_column = ['Anchorage','In Port - Shifting','Alongside Jetty','Enroute Econ Speed (85% MCR)','Enroute Econ Speed (100% MCR)','Inter Rig','Cargo Work / Passenger Transfer','Standby Close - Active in/outside 500m Zone','Standby Normal - Nil Activity','Towing @ Barge / Rig Move','Tied-up to Mooring Standby Buoy','Anchor Handling']
        TOD_column = ['No.','Name','','','','','Vessel Joining Date (DD-MM-YY)','','Length of Stay Onboard (days)','','Rank','','Nationality','','International Passport Number / NRIC','Valid Thru (Medical)','BOSIET Validity']
        rob_column = ['Open','','','Consumption','','','Loaded','','','','Discharged','','','','ROB']
        operation_act_p1_column = ['Time','','','','Activities','','','','','','','','','','','Time','','','Activities']
        operation_act_p2_column = ['Time','','Activities','','','','','','','','Time','','Activities']

        #COMBINE VERTICALLY
        combined_weather_data = weather_data_1 + weather_data_2

        #CREATE DATAFRAME
        df_summary = pd.DataFrame({'Vessel Name': [vesselName],
                           'Location': [location],
                           'Enroute to / Routing': [enroute_to],
                           'date': todayDate})
        df_weather_combine = pd.DataFrame(combined_weather_data,columns=weather_column)
        df_TOD = pd.DataFrame(TOD_data,columns=TOD_column)
        df_cargo_des = pd.DataFrame({'CARGO DESCRIPTION': cargo_descriptions})
        df_rob = pd.DataFrame(rob_data,columns=rob_column)
        df_op_act_1 = pd.DataFrame(operation_act_p1_data, columns=operation_act_p1_column)
        df_op_act_2 = pd.DataFrame(operation_act_p2_data, columns=operation_act_p2_column)
        df_activities = pd.DataFrame(activities_data,columns=activities_column)
        df_fuel = pd.DataFrame(fuel_data,columns=fuel_column)
        df_rob_combine = pd.concat((df_cargo_des,df_rob),axis=1)
        selected_fuel_row = df_fuel.iloc[day_yesterday - 1]
        selected_act_row = df_activities.iloc[day_yesterday - 1]
        selected_fuel_row_t = selected_fuel_row.to_frame().T
        selected_act_row_t = selected_act_row.to_frame().T
        #selected_row_transposed = df_activities_fuel_combine.iloc[day_yesterday - 1:day_yesterday].T

        #EXPORT TO EXCEL
        with pd.ExcelWriter(fileName, engine = 'openpyxl', mode='a') as writer:
            df_summary.to_excel(writer, sheet_name='Summary', index=False)
            df_weather_combine.to_excel(writer, sheet_name='Weather', index=False)
            selected_fuel_row_t.to_excel(writer, sheet_name='Fuel', index=False)
            selected_act_row_t.to_excel(writer, sheet_name='Activity', index=False)
            df_op_act_1.to_excel(writer, sheet_name='Op activities page 1', index=False)
            df_op_act_2.to_excel(writer, sheet_name='Op activities page 2', index=False)
            df_TOD.to_excel(writer, sheet_name='TOD', index=False)
            df_rob_combine.to_excel(writer, sheet_name='ROB', index=False)
            print(f'{vdrFile} is DONE')